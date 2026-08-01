"""
Export Excel cumulatif (onglets Afrique / Tunisie), nettoyage des
controles de formulaire residuels, validation/mise en forme conditionnelle.
Issu du decoupage de veille_ao_1_1.py (v10.18).
"""
import os
import re
import shutil
import zipfile
import hashlib
import traceback
from zipfile import ZipFile

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule

from datetime import date

from config import ENTETES_ATTENDUS, log
from utils import parse_date, normaliser_texte

_SOURCES_TUNISIE = {"TuniSurf", "TUNEPS"}

ENTETES_BASE_SANS_PDF = [e for e in ENTETES_ATTENDUS if e != "Lien PDF complet"]
ENTETES_AFRIQUE = ENTETES_ATTENDUS + ["Avis Direction Generale", "Commentaire DG", "Statut", "Jours restants"]
ENTETES_TUNISIE = ENTETES_BASE_SANS_PDF + ["Caution", "Maitre d'Ouvrage", "Avis Direction Generale", "Commentaire DG", "Statut", "Jours restants"]
_LARGEURS_BASE           = [16, 20, 15, 70, 18, 15, 55, 55]
_LARGEURS_BASE_SANS_PDF  = [16, 20, 15, 70, 18, 15, 55]
_LARGEURS_AFRIQUE = _LARGEURS_BASE + [30, 35, 18, 14]
_LARGEURS_TUNISIE = _LARGEURS_BASE_SANS_PDF + [15, 30, 30, 35, 18, 14]

_COL_ACTION_AFRIQUE = ENTETES_AFRIQUE.index("Avis Direction Generale") + 1
_COL_ACTION_TUNISIE = ENTETES_TUNISIE.index("Avis Direction Generale") + 1
_COL_STATUT_AFRIQUE = ENTETES_AFRIQUE.index("Statut") + 1
_COL_STATUT_TUNISIE = ENTETES_TUNISIE.index("Statut") + 1
_COL_JOURS_AFRIQUE  = ENTETES_AFRIQUE.index("Jours restants") + 1
_COL_JOURS_TUNISIE  = ENTETES_TUNISIE.index("Jours restants") + 1
_COL_DATE_LIMITE    = ENTETES_ATTENDUS.index("Date limite") + 1
_COL_PAYS           = ENTETES_ATTENDUS.index("Pays") + 1
_COL_SOURCE         = ENTETES_ATTENDUS.index("Source") + 1
_COL_TITRE          = ENTETES_ATTENDUS.index("Titre") + 1

_STATUT_STYLES = {
    "expire": ("🔴 Expiré",          "F7C6C6"),
    "urgent": ("🟠 Urgent (< 7j)",   "FBE5B6"),
    "ouvert": ("🟢 Ouvert",          "C6EFCE"),
    "inconnu": ("⚪ Date non précisée", "E7E6E6"),
}

_COULEUR_ACTION_CONFIRME = "57BB8A"
_COULEUR_ACTION_ATTENTE  = "FBE5B6"

def calculer_statut(date_limite_str):
    d = parse_date(date_limite_str)
    if d is None:
        return _STATUT_STYLES["inconnu"]
    jours_restants = (d - date.today()).days
    if jours_restants < 0:
        return _STATUT_STYLES["expire"]
    if jours_restants <= 7:
        return _STATUT_STYLES["urgent"]
    return _STATUT_STYLES["ouvert"]


def calculer_jours_restants(date_limite_str):
    d = parse_date(date_limite_str)
    if d is None:
        return None
    return (d - date.today()).days


_ISO2_PAR_PAYS_BRUT = {
    "Afrique du Sud": "ZA", "South Africa": "ZA",
    "Algerie": "DZ", "Algeria": "DZ",
    "Angola": "AO",
    "Benin": "BJ",
    "Botswana": "BW",
    "Burkina Faso": "BF",
    "Burundi": "BI",
    "Cameroun": "CM", "Cameroon": "CM",
    "Cap-Vert": "CV", "Cabo Verde": "CV",
    "Comores": "KM", "Comoros": "KM",
    "Congo": "CG", "Congo, Republic of": "CG",
    "Republique democratique du Congo": "CD", "Congo, Democratic Republic of": "CD",
    "Cote d'Ivoire": "CI",
    "Djibouti": "DJ",
    "Egypte": "EG", "Egypt, Arab Republic of": "EG",
    "Erythree": "ER", "Eritrea": "ER",
    "Swaziland": "SZ", "Eswatini": "SZ",
    "Ethiopie": "ET",
    "Gabon": "GA",
    "Gambie": "GM", "Gambia, The": "GM", "Gambia": "GM",
    "Ghana": "GH",
    "Guinee": "GN", "Guinea": "GN",
    "Guinee equatoriale": "GQ", "Equatorial Guinea": "GQ",
    "Guinee-Bissao": "GW", "Guinea-Bissau": "GW",
    "Kenya": "KE",
    "Lesotho": "LS",
    "Liberia": "LR",
    "Libye": "LY", "Libya": "LY",
    "Madagascar": "MG",
    "Malawi": "MW",
    "Mali": "ML",
    "Maurice": "MU", "Mauritius": "MU",
    "Mauritanie": "MR", "Mauritania": "MR",
    "Mozambique": "MZ",
    "Namibie": "NA", "Namibia": "NA",
    "Niger": "NE",
    "Nigeria": "NG",
    "Ouganda": "UG", "Uganda": "UG",
    "Republique centrafricaine": "CF", "Central African Republic": "CF",
    "Rwanda": "RW",
    "Sao Tome-et-Principe": "ST", "Sao Tome and Principe": "ST",
    "Senegal": "SN",
    "Seychelles": "SC",
    "Sierra Leone": "SL",
    "Somalie": "SO", "Somalia": "SO",
    "Soudan": "SD",
    "Sud-Soudan": "SS", "South Sudan": "SS",
    "Tanzanie": "TZ", "Tanzania": "TZ",
    "Tchad": "TD", "Chad": "TD",
    "Togo": "TG",
    "Tunisie": "TN", "Tunisia": "TN",
    "Zambie": "ZM", "Zambia": "ZM",
    "Zimbabwe": "ZW",
}


def _cle_pays(nom):
    if not nom:
        return ""
    return normaliser_texte(nom).strip().lower()


_ISO2_PAR_PAYS = {_cle_pays(k): v for k, v in _ISO2_PAR_PAYS_BRUT.items()}


def _drapeau(pays):
    code = _ISO2_PAR_PAYS.get(_cle_pays(pays))
    if not code:
        return "🌍"
    return "".join(chr(0x1F1E6 + ord(c) - ord("A")) for c in code)


def _est_source_tunisie(ao):
    if ao.get("pays", "") in ("Tunisie", "Tunisia"):
        return True
    return any(s in ao.get("source", "") for s in _SOURCES_TUNISIE)

def _preparer_feuille(wb, nom_feuille, entetes, largeurs, titre):
    if nom_feuille in wb.sheetnames:
        return wb[nom_feuille], False
    ws = wb.create_sheet(nom_feuille)
    nb_colonnes = len(entetes)

    ws.append([titre] + [""] * (nb_colonnes - 1))
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=nb_colonnes)
    cell_titre = ws.cell(row=1, column=1)
    cell_titre.fill      = PatternFill("solid", fgColor="0B2F4A")
    cell_titre.font      = Font(bold=True, color="FFFFFF", size=13)
    cell_titre.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    ws.append(entetes)
    for cell in ws[2]:
        cell.fill      = PatternFill("solid", fgColor="1F4E79")
        cell.font      = Font(bold=True, color="FFFFFF", size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for i, larg in enumerate(largeurs, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = larg
    ws.freeze_panes = "A3"
    ws.row_dimensions[2].height = 20
    return ws, True


def _recalculer_statuts_existants(ws, col_date_limite, col_statut, col_jours, col_action):
    maj = 0
    for row in ws.iter_rows(min_row=3, values_only=False):
        if row[_COL_SOURCE - 1].value is None and row[_COL_TITRE - 1].value is None:
            continue
        date_limite_brute = row[col_date_limite - 1].value or ""
        label, couleur = calculer_statut(str(date_limite_brute))
        jours = calculer_jours_restants(str(date_limite_brute))

        cell_statut = row[col_statut - 1]
        if cell_statut.value != label:
            maj += 1
        cell_statut.value      = label
        cell_statut.alignment  = Alignment(horizontal="center", vertical="center")

        cell_jours = row[col_jours - 1]
        cell_jours.value     = jours
        cell_jours.alignment = Alignment(horizontal="center", vertical="center")

        cell_date_limite = row[col_date_limite - 1]

        # Priorite : si une action DG est choisie manuellement dans Excel,
        # elle colore TOUTE la ligne (y compris Statut et Date limite),
        # recalculee a chaque export -> effective au prochain lancement.
        # Sinon, on retombe sur la couleur de statut classique
        # (expire/urgent/ouvert), appliquee seulement a Statut et Date limite.
        valeur_action = str(row[col_action - 1].value or "").strip()
        if valeur_action == "Confirmé":
            couleur_ligne = _COULEUR_ACTION_CONFIRME
        elif valeur_action == "En attente":
            couleur_ligne = _COULEUR_ACTION_ATTENTE
        else:
            couleur_ligne = None

        if couleur_ligne:
            for cell in row:
                cell.fill = PatternFill("solid", fgColor=couleur_ligne)
        else:
            cell_statut.fill      = PatternFill("solid", fgColor=couleur)
            cell_date_limite.fill = PatternFill("solid", fgColor=couleur)

    if maj:
        log.info(f"[EXCEL] Statut recalcule pour {maj} ligne(s) existante(s) dans '{ws.title}'")


def _reorganiser_par_pays(ws, nb_colonnes):
    if ws.max_row < 3:
        return

    lignes = []
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row):
        if row[_COL_SOURCE - 1].value is None and row[_COL_TITRE - 1].value is None:
            continue
        capture = [
            (c.value,
             c.fill.fgColor.rgb if c.fill and c.fill.fgColor and c.fill.fgColor.type == "rgb" else None,
             c.font.color.rgb if c.font and c.font.color and c.font.color.type == "rgb" else None,
             bool(c.font.bold) if c.font else False,
             c.font.underline if c.font else None,
             c.alignment.horizontal, c.alignment.vertical, bool(c.alignment.wrap_text), c.alignment.indent)
            for c in row
        ]
        pays = row[_COL_PAYS - 1].value or "Pays non précisé"
        lignes.append((pays, capture))

    if not lignes:
        return

    # CORRECTION : delete_rows() ne nettoie pas toujours correctement les
    # cellules fusionnees des anciens bandeaux pays, surtout apres un
    # rechargement du fichier depuis le disque (fusions fantomes qui
    # persistent sur certains numeros de ligne -> Excel affiche alors
    # la ligne comme une seule cellule fusionnee au lieu des colonnes
    # individuelles, meme si les valeurs sont bien ecrites dans le fichier).
    nb_demerges = 0
    for plage in list(ws.merged_cells.ranges):
        if plage.min_row >= 3:
            ws.unmerge_cells(str(plage))
            nb_demerges += 1
    if nb_demerges:
        log.info(f"[EXCEL] {nb_demerges} plage(s) fusionnee(s) residuelle(s) nettoyee(s) dans '{ws.title}'")

    ws.delete_rows(3, ws.max_row - 2)

    groupes = {}
    for pays, capture in lignes:
        groupes.setdefault(pays, []).append(capture)

    for pays in sorted(groupes.keys(), key=lambda p: normaliser_texte(p).lower()):
        ligne_bandeau = ws.max_row + 1
        ws.append([f"{_drapeau(pays)}  {pays}"] + [None] * (nb_colonnes - 1))
        ws.merge_cells(start_row=ligne_bandeau, start_column=1, end_row=ligne_bandeau, end_column=nb_colonnes)
        c = ws.cell(row=ligne_bandeau, column=1)
        c.fill = PatternFill("solid", fgColor="D9D9D9")
        c.font = Font(bold=True, size=11)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[ligne_bandeau].height = 20

        for capture in groupes[pays]:
            ws.append([v for v, *_ in capture])
            row_num = ws.max_row
            for i, (_, fill_rgb, font_rgb, bold, underline, horiz, vert, wrap, indent) in enumerate(capture, 1):
                cell = ws.cell(row=row_num, column=i)
                if fill_rgb:
                    cell.fill = PatternFill("solid", fgColor=fill_rgb)
                if font_rgb or bold or underline:
                    cell.font = Font(color=font_rgb, bold=bold, underline=underline)
                cell.alignment = Alignment(horizontal=horiz, vertical=vert, wrap_text=wrap, indent=indent)


def _lire_ids_existants(ws, colonne_lien_avis):
    ids = set()
    _RE_HYPERLINK_FORMULE = re.compile(r'^=HYPERLINK\("([^"]+)"', re.IGNORECASE)
    idx_lien = colonne_lien_avis - 1
    for row in ws.iter_rows(min_row=2, values_only=False):
        source_cell = row[1].value or ""
        titre_cell  = row[3].value or ""
        lien_tag  = row[idx_lien].hyperlink
        if lien_tag:
            lien_cell = lien_tag.target
        else:
            val_brut = str(row[idx_lien].value or "")
            m = _RE_HYPERLINK_FORMULE.match(val_brut)
            lien_cell = m.group(1).replace('""', '"') if m else ""
        cle = hashlib.md5((titre_cell + lien_cell + source_cell).encode("utf-8")).hexdigest()
        ids.add(cle)
    return ids


def _ajouter_validation_action(ws, colonne_action, derniere_ligne=5000):
    dv = DataValidation(type="list", formula1='"Confirmé,En attente,Masquer"', allow_blank=True)
    ws.add_data_validation(dv)
    lettre = openpyxl.utils.get_column_letter(colonne_action)
    plage = f"{lettre}3:{lettre}{derniere_ligne}"
    dv.add(plage)

    ws.conditional_formatting.add(plage, CellIsRule(operator="equal", formula=['"Confirmé"'],
        fill=PatternFill("solid", fgColor="C6EFCE")))
    ws.conditional_formatting.add(plage, CellIsRule(operator="equal", formula=['"En attente"'],
        fill=PatternFill("solid", fgColor="FFF2CC")))
    ws.conditional_formatting.add(plage, CellIsRule(operator="equal", formula=['"Masquer"'],
        fill=PatternFill("solid", fgColor="D9D9D9")))


_COL_COMMENTAIRE_AFRIQUE = ENTETES_AFRIQUE.index("Commentaire DG") + 1
_COL_COMMENTAIRE_TUNISIE = ENTETES_TUNISIE.index("Commentaire DG") + 1


ENTETES_ARCHIVES = [
    "Onglet origine", "Date ajout", "Source", "Pays", "Titre",
    "Date publication", "Date limite", "Lien avis",
    "Avis Direction Generale", "Commentaire DG",
    "Statut (à l'archivage)", "Date archivage",
]
_LARGEURS_ARCHIVES = [14, 14, 20, 15, 60, 18, 15, 40, 18, 35, 20, 14]
_RE_HYPERLINK_ARCHIVE = re.compile(r'^=HYPERLINK\("([^"]+)"', re.IGNORECASE)


def _preparer_feuille_archives(wb):
    if "Archives" in wb.sheetnames:
        return wb["Archives"]
    ws = wb.create_sheet("Archives")
    nb = len(ENTETES_ARCHIVES)
    ws.append(["AO masqués — archivés automatiquement, jamais réinsérés dans Afrique/Tunisie"] + [""] * (nb - 1))
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=nb)
    c = ws.cell(row=1, column=1)
    c.fill = PatternFill("solid", fgColor="595959")
    c.font = Font(bold=True, color="FFFFFF", size=13)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    ws.append(ENTETES_ARCHIVES)
    for cell in ws[2]:
        cell.fill = PatternFill("solid", fgColor="595959")
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for i, larg in enumerate(_LARGEURS_ARCHIVES, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = larg
    ws.freeze_panes = "A3"
    ws.row_dimensions[2].height = 20
    return ws


def _lire_ids_archives(ws_archives):
    ids = set()
    if ws_archives is None:
        return ids
    idx_source, idx_titre, idx_lien = 2, 4, 7
    for row in ws_archives.iter_rows(min_row=3, values_only=False):
        if row[idx_source].value is None and row[idx_titre].value is None:
            continue
        source = row[idx_source].value or ""
        titre  = row[idx_titre].value or ""
        val_brut = str(row[idx_lien].value or "")
        m = _RE_HYPERLINK_ARCHIVE.match(val_brut)
        lien = m.group(1).replace('""', '"') if m else val_brut
        cle = hashlib.md5((titre + lien + source).encode("utf-8")).hexdigest()
        ids.add(cle)
    return ids


def _archiver_lignes_masquees(ws_archives, ws, nom_onglet, col_action, col_commentaire, col_statut):
    lignes_a_supprimer = []
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=False):
        if row[_COL_SOURCE - 1].value is None and row[_COL_TITRE - 1].value is None:
            continue
        valeur_action = str(row[col_action - 1].value or "").strip()
        if valeur_action != "Masquer":
            continue

        ws_archives.append([
            nom_onglet,
            row[0].value,
            row[_COL_SOURCE - 1].value,
            row[_COL_PAYS - 1].value,
            row[_COL_TITRE - 1].value,
            row[4].value,
            row[_COL_DATE_LIMITE - 1].value,
            row[6].value,
            row[col_action - 1].value,
            row[col_commentaire - 1].value,
            row[col_statut - 1].value,
            date.today().strftime("%d/%m/%Y"),
        ])
        ligne_archivee = ws_archives.max_row
        for cell in ws_archives[ligne_archivee]:
            cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell_lien = ws_archives.cell(row=ligne_archivee, column=8)
        if isinstance(cell_lien.value, str) and cell_lien.value.startswith("=HYPERLINK"):
            cell_lien.font = Font(color="0563C1", underline="single")

        lignes_a_supprimer.append(row[0].row)

    for row_idx in sorted(lignes_a_supprimer, reverse=True):
        ws.delete_rows(row_idx)

    if lignes_a_supprimer:
        log.info(f"[EXCEL] {len(lignes_a_supprimer)} ligne(s) 'Masquer' archivee(s) depuis '{nom_onglet}' -> onglet Archives")


_COULEURS_SOURCES = {
    "SBEE": "D9F2D9",
    "AfDB": "D6E4F7",
    "DGCMEF": "F7D9D9",
    "World Bank": "D6F0F5",
    "OPEC Fund": "FBF0D0",
    "IsDB": "D3F0E0",
    "DGMarket": "F5D9EA",
    "TUNEPS": "F7E0D0",
    "TuniSurf": "E6D9F5",
    "DevelopmentAid": "FFE0B2",
    "GlobalTenders": "D6D9F7",
    "J360": "FFE8F0",
}


def _creer_feuille_legende(wb):
    if "Légende" in wb.sheetnames:
        return
    ws = wb.create_sheet("Légende")
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 70

    def titre(texte, ligne):
        ws.cell(row=ligne, column=1, value=texte)
        ws.merge_cells(start_row=ligne, start_column=1, end_row=ligne, end_column=2)
        c = ws.cell(row=ligne, column=1)
        c.fill = PatternFill("solid", fgColor="0B2F4A")
        c.font = Font(bold=True, color="FFFFFF", size=12)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[ligne].height = 22

    def paire(label, couleur, explication, ligne):
        c1 = ws.cell(row=ligne, column=1, value=label)
        c1.fill = PatternFill("solid", fgColor=couleur)
        c1.alignment = Alignment(horizontal="center", vertical="center")
        c1.font = Font(bold=True)
        c2 = ws.cell(row=ligne, column=2, value=explication)
        c2.alignment = Alignment(vertical="center", wrap_text=True)

    ligne = 1
    titre("Statut (colonne auto-calculée à chaque export)", ligne); ligne += 1
    for label, couleur in _STATUT_STYLES.values():
        explications = {
            "🔴 Expiré": "Date limite dépassée — l'AO reste visible pour l'historique mais n'est plus à traiter.",
            "🟠 Urgent (< 7j)": "Date limite dans 7 jours ou moins — à traiter en priorité.",
            "🟢 Ouvert": "Date limite encore éloignée (> 7 jours), ou pas de date limite mais AO actif.",
            "⚪ Date non précisée": "Impossible de déterminer la date limite — à vérifier manuellement.",
        }
        paire(label, couleur, explications[label], ligne)
        ligne += 1
    ligne += 1

    titre("Colonne \"Avis Direction Generale\" (action manuelle)", ligne); ligne += 1
    paire("Confirmé",   "C6EFCE", "AO validé par la Direction Générale.", ligne); ligne += 1
    paire("En attente", "FFF2CC", "AO soumis à la DG, décision en cours.", ligne); ligne += 1
    paire("Masquer",    "D9D9D9", "AO non pertinent, masqué sans suppression.", ligne)


def nettoyer_controles_formulaire(path):
    if not os.path.exists(path):
        return

    try:
        with zipfile.ZipFile(path, "r") as z:
            noms = z.namelist()

        fichiers_controles = [
            n for n in noms
            if re.search(r"xl/ctrlProps/ctrlProp\d+\.xml$", n)
            or re.search(r"xl/drawings/vmlDrawing\d+\.vml$", n)
        ]

        if not fichiers_controles:
            log.debug(f"[NETTOYAGE CONTROLES] Aucun controle de formulaire trouve dans {path}")
            return

        log.info(f"[NETTOYAGE CONTROLES] {len(fichiers_controles)} fichier(s) de controle trouve(s) -> suppression")

        path_tmp = path + ".tmp"
        with zipfile.ZipFile(path, "r") as z_in:
            # CORRECTION : [Content_Types].xml doit etre le premier fichier
            # dans le ZIP sinon Excel corrompt le fichier a l'ouverture.
            tous_items = sorted(
                z_in.infolist(),
                key=lambda x: (x.filename != "[Content_Types].xml", x.filename)
            )
            with zipfile.ZipFile(path_tmp, "w", zipfile.ZIP_DEFLATED) as z_out:
                for item in tous_items:
                    if item.filename in fichiers_controles:
                        continue
                    z_out.writestr(item, z_in.read(item.filename))

        shutil.move(path_tmp, path)
        log.info(f"[NETTOYAGE CONTROLES] Nettoyage termine avec succes sur {path}")

    except Exception as e:
        log.warning(f"[NETTOYAGE CONTROLES] Erreur lors du nettoyage (ignoree, sans impact) : {e}")


def exporter_excel(nouveaux, path):
    log.info(f"[EXCEL] Debut export : {len(nouveaux)} AO actifs a traiter -> fichier {path}")
    if not nouveaux and not os.path.exists(path):
        log.info("[EXCEL] Aucun AO et aucun fichier existant -> rien a faire")
        return
    if not nouveaux:
        log.info("[EXCEL] Aucun nouvel AO, mais recalcul du Statut des lignes existantes")
    try:
        fichier_existe = os.path.exists(path)

        if fichier_existe:
            wb_test = openpyxl.load_workbook(path)
            if "Afrique" not in wb_test.sheetnames and "Tunisie" not in wb_test.sheetnames:
                backup_path = path.replace(".xlsx", "_ancien_format.xlsx.bak")
                log.warning(f"[EXCEL] Ancien format detecte -> archivage vers {backup_path}")
                os.replace(path, backup_path)
                fichier_existe = False

        if fichier_existe:
            wb = openpyxl.load_workbook(path)
            log.info(f"[EXCEL] Fichier existant charge : {path}")

        else:
            log.info("[EXCEL] Creation d'un nouveau classeur (2 onglets)")
            wb = openpyxl.Workbook()
            wb.remove(wb.active)

        ws_afrique, cree_afrique = _preparer_feuille(
            wb, "Afrique", ENTETES_AFRIQUE, _LARGEURS_AFRIQUE,
            "Etat de suivi veille commerciale et annonce appels d'offre en Afrique",
        )
        ws_tunisie, cree_tunisie = _preparer_feuille(
            wb, "Tunisie", ENTETES_TUNISIE, _LARGEURS_TUNISIE,
            "Etat de suivi veille commerciale et annonce appels d'offre en Tunisie",
        )
        if cree_afrique:
            _ajouter_validation_action(ws_afrique, _COL_ACTION_AFRIQUE)
        if cree_tunisie:
            _ajouter_validation_action(ws_tunisie, _COL_ACTION_TUNISIE)

        _creer_feuille_legende(wb)

        ws_archives = _preparer_feuille_archives(wb)
        _archiver_lignes_masquees(
            ws_archives, ws_afrique, "Afrique",
            _COL_ACTION_AFRIQUE, _COL_COMMENTAIRE_AFRIQUE, _COL_STATUT_AFRIQUE,
        )
        _archiver_lignes_masquees(
            ws_archives, ws_tunisie, "Tunisie",
            _COL_ACTION_TUNISIE, _COL_COMMENTAIRE_TUNISIE, _COL_STATUT_TUNISIE,
        )

        ids_archives = _lire_ids_archives(ws_archives)
        ids_afrique = _lire_ids_existants(ws_afrique, colonne_lien_avis=7) | ids_archives
        ids_tunisie = _lire_ids_existants(ws_tunisie, colonne_lien_avis=7) | ids_archives
        log.info(
            f"[EXCEL] Lignes deja presentes -> Afrique={len(ids_afrique)} | Tunisie={len(ids_tunisie)} "
            f"| Archives (masques, exclus definitivement)={len(ids_archives)}"
        )

        _recalculer_statuts_existants(ws_afrique, _COL_DATE_LIMITE, _COL_STATUT_AFRIQUE, _COL_JOURS_AFRIQUE, _COL_ACTION_AFRIQUE)
        _recalculer_statuts_existants(ws_tunisie, _COL_DATE_LIMITE, _COL_STATUT_TUNISIE, _COL_JOURS_TUNISIE, _COL_ACTION_TUNISIE)

        couleurs = {}

        ajoutes = {"Afrique": 0, "Tunisie": 0}
        deja_presents = {"Afrique": 0, "Tunisie": 0}

        for ao in nouveaux:
            source      = ao.get("source", "")
            titre       = ao.get("titre", "")
            url_avis_id = ao.get("url_avis", "")
            url_dossier = ao.get("url_dossier", "")
            cle_ao = hashlib.md5((titre + url_avis_id + source).encode("utf-8")).hexdigest()

            est_tunisie = _est_source_tunisie(ao)
            ws         = ws_tunisie if est_tunisie else ws_afrique
            ids_cibles = ids_tunisie if est_tunisie else ids_afrique
            nom_onglet = "Tunisie" if est_tunisie else "Afrique"

            if cle_ao in ids_cibles:
                deja_presents[nom_onglet] += 1
                continue

            couleur = "FFFFFF"
            page    = ao.get("page_pdf", "")

            if est_tunisie:
                ligne = [
                    ao.get("date_ajout", ""), source, ao.get("pays", ""),
                    titre, ao.get("date_pub", ""), ao.get("date_limite", ""),
                    url_avis_id,
                    ao.get("caution", ""), ao.get("acheteur_public", ""), "", "",
                ]
                col_lien_avis = 7
                col_pdf       = None
            else:
                ligne = [
                    ao.get("date_ajout", ""), source, ao.get("pays", ""),
                    titre, ao.get("date_pub", ""), ao.get("date_limite", ""),
                    url_avis_id, url_dossier,
                ] + ["", ""]
                col_lien_avis = 7
                col_pdf       = 8

            ws.append(ligne)
            row  = ws.max_row
            fill = PatternFill("solid", fgColor=couleur)
            for cell in ws[row]:
                cell.fill      = fill
                cell.alignment = Alignment(vertical="center", wrap_text=True)

            if url_avis_id:
                cell = ws.cell(row=row, column=col_lien_avis)

                # ── Cas particulier GlobalTenders ──────────────────────────
                # Le lien du site exige une session connectee (voir capture
                # d'ecran : redirection vers le login). Un PDF local a ete
                # genere pendant le scraping (page.pdf() via session active,
                # cf. globaltenders.py / champ "pdf_local"). On pointe donc
                # le lien Excel vers ce fichier local, en chemin RELATIF au
                # dossier du fichier Excel, pour que le lien reste valide
                # meme si on change de PC/environnement — a condition de
                # garder le dossier PDF a cote du fichier Excel (ideal :
                # dossier synchronise OneDrive/Google Drive).
                pdf_local = ao.get("pdf_local", "")
                if "GlobalTenders" in source and pdf_local and os.path.exists(pdf_local):
                    dossier_excel = os.path.dirname(os.path.abspath(path)) or "."
                    chemin_relatif = os.path.relpath(os.path.abspath(pdf_local), start=dossier_excel)
                    chemin_relatif = chemin_relatif.replace("\\", "/")
                    cible_lien = chemin_relatif
                    label = "Ouvrir l'avis (PDF local)"
                elif ao.get("lien_generique") and "TuniSurf" in source:
                    cible_lien = url_avis_id
                    label = "⚠️ Chercher manuellement sur TuniSurf (N° " + ao.get("reference", "") + ")"
                elif "TuniSurf" in source:
                    cible_lien = url_avis_id
                    label = "Ouvrir l'avis (necessite connexion TuniSurf)"
                else:
                    cible_lien = url_avis_id
                    label = f"-> page {page}" if page else "Ouvrir l'avis"

                url_formule = cible_lien.replace('"', '""')
                cell.value  = f'=HYPERLINK("{url_formule}","{label}")'
                cell.font   = Font(color="0563C1", underline="single", bold=bool(page))

            if col_pdf and url_dossier and url_dossier != url_avis_id:
                cell        = ws.cell(row=row, column=col_pdf)
                url_formule = url_dossier.replace('"', '""')
                cell.value  = f'=HYPERLINK("{url_formule}","PDF complet")'
                cell.font   = Font(color="0563C1", underline="single")

            col_statut = _COL_STATUT_TUNISIE if est_tunisie else _COL_STATUT_AFRIQUE
            col_jours  = _COL_JOURS_TUNISIE if est_tunisie else _COL_JOURS_AFRIQUE
            label_statut, couleur_statut = calculer_statut(ao.get("date_limite", ""))
            jours_restants = calculer_jours_restants(ao.get("date_limite", ""))

            cell_statut = ws.cell(row=row, column=col_statut)
            cell_statut.value      = label_statut
            cell_statut.fill       = PatternFill("solid", fgColor=couleur_statut)
            cell_statut.alignment  = Alignment(horizontal="center", vertical="center")

            cell_jours = ws.cell(row=row, column=col_jours)
            cell_jours.value      = jours_restants
            cell_jours.alignment  = Alignment(horizontal="center", vertical="center")

            ws.cell(row=row, column=_COL_DATE_LIMITE).fill = PatternFill("solid", fgColor=couleur_statut)

            ids_cibles.add(cle_ao)
            ajoutes[nom_onglet] += 1

        _reorganiser_par_pays(ws_afrique, len(ENTETES_AFRIQUE))

        derniere_col_afrique = openpyxl.utils.get_column_letter(len(ENTETES_AFRIQUE))
        derniere_col_tunisie = openpyxl.utils.get_column_letter(len(ENTETES_TUNISIE))
        ws_afrique.auto_filter.ref = f"A2:{derniere_col_afrique}{ws_afrique.max_row}"
        ws_tunisie.auto_filter.ref = f"A2:{derniere_col_tunisie}{ws_tunisie.max_row}"

        wb.save(path)
        nettoyer_controles_formulaire(path)
        log.info(
            f"[EXCEL] Mise a jour terminee : {path} | "
            f"Afrique : {ajoutes['Afrique']} ajoute(s), {deja_presents['Afrique']} deja presente(s) | "
            f"Tunisie : {ajoutes['Tunisie']} ajoute(s), {deja_presents['Tunisie']} deja presente(s)"
        )
    except Exception as e:
        log.error(f"[EXCEL] Erreur export Excel : {e}")
        log.error(traceback.format_exc())